import csv
import argparse
import openpyxl
import sys

# Get headers from CSV file
def get_headers_csv(file_name):
    try:
        with open(file_name, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            headers = next(reader)
        return headers
    except FileNotFoundError:
        print(f"Error: File '{file_name}' not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

# Get headers from Excel file
def get_headers_excel(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows())]
        return headers
    except FileNotFoundError:
        print(f"Error: File '{file_name}' not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

# Display headers
def display_headers(headers):
    print("Available headers:")
    for i, header in enumerate(headers, 1):
        print(f"{i}. {header}")

# Read column data from CSV file
def read_column_csv(file_name, header):
    try:
        with open(file_name, 'r', newline='') as csvfile:
            reader = csv.DictReader(csvfile)
            print(f"\nData for column '{header}':")
            for row in reader:
                print(row[header])
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Read column data from Excel file
def read_column_excel(file_name, header):
    try:
        workbook = openpyxl.load_workbook(file_name, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows())]
        col_index = headers.index(header) + 1
        print(f"\nData for column '{header}':")
        for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
            print(row[0].value)
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Main program
def main():
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description="Extract column data from CSV or Excel files.")
    group = parser.add_mutually_exclusive_group(required=True)
    
    # Add mutually exclusive arguments for CSV and Excel files
    group.add_argument("-c", "--csv", help="Input CSV file")
    group.add_argument("-e", "--excel", help="Input Excel file")
    args = parser.parse_args()

    if args.csv:
        file_name = args.csv
        get_headers = get_headers_csv
        read_column = read_column_csv
    else:
        file_name = args.excel
        get_headers = get_headers_excel
        read_column = read_column_excel

    headers = get_headers(file_name)

    # Check if headers were retrieved successfully
    if headers:
        display_headers(headers)
        
        # Prompt user to select a header
        while True:
            try:
                choice = int(input("\nEnter the number of the header you want to extract (1-{}): ".format(len(headers))))
                if 1 <= choice <= len(headers):
                    selected_header = headers[choice - 1]
                    read_column(file_name, selected_header)
                    break
                else:
                    print("Invalid number. Please try again.")
            except ValueError:
                print("Please enter a valid number.")

if __name__ == "__main__":
    main()
